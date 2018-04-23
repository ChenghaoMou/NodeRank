#!/usr/bin/env python
# -*- coding: utf-8 -*-
# @Time    : 17-12-12 下午7:58
# @Author  : Aries
# @Site    : 
# @File    : Algorithm.py
# @Software: PyCharm
from __future__ import division
from __future__ import print_function

from collections import OrderedDict, defaultdict
from pprint import pprint

import networkx as nx
import numpy
import pandas as pd
from gensim.models.keyedvectors import KeyedVectors
from tqdm import tqdm
from openpyxl import Workbook

from Preprocess.Constructors import NetworkConstructor, group_list, interval_constructor
from Preprocess.Conf import path_constructor, configurations, running_configurations


def cache(func):
    __memory__ = dict()

    def wrapper(self, *args, **kwargs):
        if func not in __memory__.keys():
            __memory__[func.func_name] = defaultdict(float)
            __memory__[func.func_name][args] = func(self, *args, **kwargs)
        return __memory__[func.func_name][args]

    return wrapper


# pd.set_option("display.precision", 10)

class NodeRank(object):

    def __init__(self, dissemination_network=None):

        self.node_tables = dict()

        # self.relation_graph = relation_network
        self.dissemination_graph = dissemination_network
        self.node_influence = pd.Series(0.0, index=list(sorted(self.dissemination_graph.nodes())))
        self.node_vectors = KeyedVectors.load_word2vec_format('/home/maxen/Documents/Code/PycharmProjects/NodeRank/data/higgs/higgs.emb', binary=False)

    def initialize(self):
        for node in tqdm(self.dissemination_graph.nodes()):
            in_edges = map(lambda x: ["In"] + list(x), list(self.dissemination_graph.in_edges(node, data=True)))
            out_edges = map(lambda x: ["Out"] + list(x), list(self.dissemination_graph.out_edges(node, data=True)))

            ## Edge ('Direction(In|Out)', source, target, {"time":0, "type":"RT", "weight":0.0})

            intervals, column_index = interval_constructor(
                group_list(in_edges + out_edges, group_by=lambda x: x[-1]['time'], squeeze=True, squeeze_by=lambda x: x[0]), group_by=lambda x:x[0])

            row_index = list(self.dissemination_graph.successors(node))

            f = pd.DataFrame(0.0, index=row_index, columns=column_index, dtype=numpy.float64)

            for interval, edges in intervals.items():
                f.loc[[edge[2] for edge in edges["Out"]], interval] = 1.0

            f = f.div(f.sum(axis=0)).fillna(0.0)

            p = pd.Series(1.0 / nx.number_of_nodes(self.dissemination_graph) / len(row_index) if row_index else 0.0,
                          index=row_index, dtype=numpy.float64)
            c = pd.Series(0.0, index=column_index, dtype=numpy.float64)

            self.node_tables[node] = {
                "Intervals": intervals,
                "Index": column_index,
                "P": p,
                "F": f,
                "C": c
            }

            # print(node)
            # print(p)

    def run(self, iteration=2, epsilon=1e-10, damping=0.85, ignore_resemblance=True, ignore_action_type=True):

        previous = self.node_influence.copy()

        for i in range(iteration):
            for node in self.dissemination_graph.nodes():
                if not self.node_tables[node]["C"].empty: self.collect(node, damping=damping,
                                                                       ignore_resemblance=ignore_resemblance,
                                                                       ignore_action_type=ignore_action_type)

            for node in self.dissemination_graph.nodes():
                if not self.node_tables[node]["P"].empty: self.vote(node)
                self.node_influence[node] = self.node_tables[node]["C"].sum()


            err = sum((self.node_influence - previous).abs())

            if err < nx.number_of_nodes(self.dissemination_graph) * epsilon:
                print(
                    "Finished {1}-th iteration with epsilon {0:.20f}".format(err,
                                                                            i))
                break
            else:
                print(
                    "Continue {1}-th iteration with epsilon {0:.20f}".format(err,
                                                                            i), end='\r')
                previous = self.node_influence.copy()

    def collect(self, target_node, damping=0.85, ignore_resemblance=True, ignore_action_type=True):
        temp = (1 - damping) / nx.number_of_nodes(self.dissemination_graph)
        for interval, edges in self.node_tables[target_node]["Intervals"].items():
            self.node_tables[target_node]["C"].ix[
                self.node_tables[target_node]["Index"].index(interval), 0] = temp + damping * sum(
                self.node_tables[edge[1]]["P"][edge[2]] *
                (1 if ignore_resemblance else self.resemblance(target_node, edge[1], alpha=0.8)) *
                (1 if ignore_action_type else edge[-1]["weight"])
                for edge in edges["In"])

        return

    @cache
    def resemblance(self, reference_node, neighbor, alpha=0.5):

        return self.node_vectors.similarity(str(reference_node), str(neighbor))

    def vote(self, node):
        self.node_tables[node]["P"] = pd.Series(data=self.node_tables[node]["F"].dot(self.node_tables[node]["C"]),
                                                index=self.node_tables[node]["P"].index,
                                                dtype=numpy.float64)

    def sort(self):
        temp = defaultdict(list)
        for pair in zip(list(self.node_influence.index), list(self.node_influence.values)):
            temp["{0:.20f}".format(pair[-1])].append(pair)
        for key in temp.keys():
            temp[key] = sorted(temp[key], key=lambda x:x[1], reverse=False)

        result = OrderedDict()
        for key, value in sorted(temp.items(), key=lambda x:x[0], reverse=True):
            for pair in value:
                result[pair[0]] = pair[1]
        # print(result)
        self.node_influence = result

    @cache
    def num_valid_contributor(self, target_node, seed, depth):
        if not seed:
            seed = 0

        result = 0
        for edge in self.dissemination_graph.in_edges(target_node, data=True):
            if depth < 5 and edge[-1]['time'] >= seed: result += 1 + self.num_valid_contributor(edge[0], edge[-1]['time'], depth+1)

        return result

    def num_valid_collector(self, target_node, seed, depth):
        if not seed: seed = max([e[-1]['time'] for e in self.dissemination_graph.in_edges(target_node, data=True)])

        result = 0
        for edge in self.dissemination_graph.out_edges(target_node, data=True):
            if depth < 1 and edge[-1]["time"] <= seed: result += 1 + self.num_valid_collector(edge[1], edge[-1]['time'], depth+1)

        return result


# def m_pagerank(G, alpha=0.85, personalization=None,
#              max_iter=100, tol=1.0e-6, nstart=None, weight='weight',
#              dangling=None):
#
#     """Return the PageRank of the nodes in the graph.
#
#     PageRank computes a ranking of the nodes in the graph G based on
#     the structure of the incoming links. It was originally designed as
#     an algorithm to rank web pages.
#
#     Parameters
#     ----------
#     G : graph
#       A NetworkX graph.  Undirected graphs will be converted to a directed
#       graph with two directed edges for each undirected edge.
#
#     alpha : float, optional
#       Damping parameter for PageRank, default=0.85.
#
#     personalization: dict, optional
#       The "personalization vector" consisting of a dictionary with a
#       key some subset of graph nodes and personalization value each of those.
#       At least one personalization value must be non-zero.
#       If not specfiied, a nodes personalization value will be zero.
#       By default, a uniform distribution is used.
#
#     max_iter : integer, optional
#       Maximum number of iterations in power method eigenvalue solver.
#
#     tol : float, optional
#       Error tolerance used to check convergence in power method solver.
#
#     nstart : dictionary, optional
#       Starting value of PageRank iteration for each node.
#
#     weight : key, optional
#       Edge data key to use as weight.  If None weights are set to 1.
#
#     dangling: dict, optional
#       The outedges to be assigned to any "dangling" nodes, i.e., nodes without
#       any outedges. The dict key is the node the outedge points to and the dict
#       value is the weight of that outedge. By default, dangling nodes are given
#       outedges according to the personalization vector (uniform if not
#       specified). This must be selected to result in an irreducible transition
#       matrix (see notes under google_matrix). It may be common to have the
#       dangling dict to be the same as the personalization dict.
#
#     Returns
#     -------
#     pagerank : dictionary
#        Dictionary of nodes with PageRank as value
#
#     Examples
#     --------
#     >>> G = nx.DiGraph(nx.path_graph(4))
#     >>> pr = nx.pagerank(G, alpha=0.9)
#
#     Notes
#     -----
#     The eigenvector calculation is done by the power iteration method
#     and has no guarantee of convergence.  The iteration will stop after
#     an error tolerance of ``len(G) * tol`` has been reached. If the
#     number of iterations exceed `max_iter`, a
#     :exc:`networkx.exception.PowerIterationFailedConvergence` exception
#     is raised.
#
#     The PageRank algorithm was designed for directed graphs but this
#     algorithm does not check if the input graph is directed and will
#     execute on undirected graphs by converting each edge in the
#     directed graph to two edges.
#
#     See Also
#     --------
#     pagerank_numpy, pagerank_scipy, google_matrix
#
#     Raises
#     ------
#     PowerIterationFailedConvergence
#         If the algorithm fails to converge to the specified tolerance
#         within the specified number of iterations of the power iteration
#         method.
#
#     References
#     ----------
#     .. [1] A. Langville and C. Meyer,
#        "A survey of eigenvector methods of web information retrieval."
#        http://citeseer.ist.psu.edu/713792.html
#     .. [2] Page, Lawrence; Brin, Sergey; Motwani, Rajeev and Winograd, Terry,
#        The PageRank citation ranking: Bringing order to the Web. 1999
#        http://dbpubs.stanford.edu:8090/pub/showDoc.Fulltext?lang=en&doc=1999-66&format=pdf
#
#     """
#     if len(G) == 0:
#         return {}
#
#     if not G.is_directed():
#         D = G.to_directed()
#     else:
#         D = G
#
#     # Create a copy in (right) stochastic form
#     W = nx.stochastic_graph(D, weight=weight)
#     N = W.number_of_nodes()
#
#     # Choose fixed starting vector if not given
#     if nstart is None:
#         x = dict.fromkeys(W, 1.0 / N)
#     else:
#         # Normalized nstart vector
#         s = float(sum(nstart.values()))
#         x = dict((k, v / s) for k, v in nstart.items())
#
#     if personalization is None:
#         # Assign uniform personalization vector if not given
#         p = dict.fromkeys(W, 1.0 / N)
#     else:
#         s = float(sum(personalization.values()))
#         p = dict((k, v / s) for k, v in personalization.items())
#
#     if dangling is None:
#         # Use personalization vector if dangling vector not specified
#         dangling_weights = p
#     else:
#         s = float(sum(dangling.values()))
#         dangling_weights = dict((k, v/s) for k, v in dangling.items())
#     dangling_nodes = [n for n in W if W.out_degree(n, weight=weight) == 0.0]
#
#     # power iteration: make up to max_iter iterations
#     for _ in range(max_iter):
#         xlast = x
#         x = dict.fromkeys(xlast.keys(), 0)
#         danglesum = alpha * sum(xlast[n] for n in dangling_nodes)
#         for n in x:
#             # this matrix multiply looks odd because it is
#             # doing a left multiply x^T=xlast^T*W
#             for nbr in W[n]:
#                 x[nbr] += alpha * xlast[n] * W[n][nbr][weight]
#             x[n] += (1.0 - alpha) * p.get(n,0)
#         # check convergence, l1 norm
#         err = sum([abs(x[n] - xlast[n]) for n in x])
#         print("{0}-th iteration with {1:.10f}".format(_, err), end="\r")
#         if err < N*tol:
#             return x
#     raise nx.PowerIterationFailedConvergence(max_iter)


if __name__ == "__main__":

    import pickle
    idx2word = pickle.load(open("/home/maxen/Documents/Code/PycharmProjects/NodeRank/Preprocess/wordindex.dict"))

    dissemination_graph = NetworkConstructor(**running_configurations["text"]).construct()

    rank = NodeRank(dissemination_network=dissemination_graph)
    rank.initialize()
    rank.run(iteration=200, damping=0.85, epsilon=1e-6, ignore_action_type=False, ignore_resemblance=True)
    rank.sort()

    page_ranks = nx.pagerank(dissemination_graph, alpha=0.85, tol=1e-6, max_iter=200, weight=None)
    page_ranks_index = OrderedDict()

    for index, node in enumerate(sorted(page_ranks.items(), key=lambda x: x[1], reverse=True)):
        page_ranks_index[node[0]] = (index, node[1])

    average_shift = 0.0

    # region Write to an excel file.
    wb = Workbook()
    output_file = path_constructor('data/output/word.xlsx')

    worksheet = wb.active
    worksheet.title = "Ranking comparison"

    for idx, (node, value) in enumerate(rank.node_influence.iteritems()):
        # average_shift += abs(page_ranks_index[node][0] - idx)

        worksheet.cell(column=1, row=idx + 1, value="{0:<10}".format(idx2word.get(node).encode("utf-8")))
        worksheet.cell(column=2, row=idx + 1, value="{0:<10}".format(idx))
        worksheet.cell(column=3, row=idx + 1, value="{0:<10,.10f}".format(value))
        worksheet.cell(column=4, row=idx + 1, value="{0:<10}".format(page_ranks_index[node][0]))
        worksheet.cell(column=5, row=idx + 1, value="{0:<10}".format(page_ranks_index[node][1]))
        worksheet.cell(column=6, row=idx + 1, value="{0:<10}".format(("↑" if page_ranks_index[node][0] > idx else "↓") + str(abs(page_ranks_index[node][0] - idx))))

    wb.save(filename=output_file)

    # print("Average shift: {0:.2f}".format(average_shift/len(rank.node_influence)))

    # endregion
